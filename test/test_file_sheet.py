"""Tests for /api/file-sheet — xlsx parsing, formula fallback, caps, and denials."""

from __future__ import annotations

import asyncio
import io
from unittest.mock import MagicMock, patch

import pytest
from aiohttp import web
from aiohttp.test_utils import TestClient, TestServer

from kiro_crew.dashboard.handlers import api_file_sheet
from kiro_crew.dashboard.handlers.files import (
    _SHEET_MAX_COLS,
    _SHEET_MAX_ROWS,
    _load_sheet_payload,
    _parse_workbook_grid,
    _sheet_cell_json,
    _SheetRefusal,
)


def _make_app() -> web.Application:
    app = web.Application()
    app.router.add_get("/api/file-sheet", api_file_sheet)
    return app


@pytest.fixture
def mock_sel():
    with patch("kiro_crew.sel.sel") as m, \
         patch("kiro_crew.security.is_sensitive_path", return_value=False):
        instance = MagicMock()
        m.return_value = instance
        yield instance


def _workbook_bytes(populate) -> bytes:
    """Build an in-memory xlsx via openpyxl and hand back the raw bytes."""
    from openpyxl import Workbook

    wb = Workbook()
    populate(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Happy path -------------------------------------------------------------


@pytest.mark.asyncio
async def test_returns_cell_grid_with_sheet_names(tmp_path, mock_sel):
    def populate(wb):
        ws = wb.active
        ws.title = "DCF"
        ws["A1"] = "Revenue"
        ws["B1"] = 1000
        ws["A2"] = "Margin"
        ws["B2"] = 0.42
        ws2 = wb.create_sheet("Assumptions")
        ws2["A1"] = "WACC"

    f = tmp_path / "model.xlsx"
    f.write_bytes(_workbook_bytes(populate))
    with patch("kiro_crew.dashboard.handlers._validate_dashboard_path", return_value=str(f)):
        async with TestClient(TestServer(_make_app())) as client:
            resp = await client.get(f"/api/file-sheet?path={f}")
            assert resp.status == 200
            body = await resp.json()
    assert [s["name"] for s in body["sheets"]] == ["DCF", "Assumptions"]
    assert body["sheets"][0]["rows"] == [["Revenue", 1000], ["Margin", 0.42]]
    assert body["sheets"][1]["rows"] == [["WACC"]]
    assert body["total_sheets"] == 2
    assert body["truncated_sheets"] is False


def test_formula_without_cached_value_falls_back_to_formula_text():
    # openpyxl-generated workbooks carry no cached results, which is exactly
    # the shape agent-produced spreadsheets have. The grid must show the
    # formula source rather than an empty cell.
    def populate(wb):
        ws = wb.active
        ws["A1"] = 5
        ws["B1"] = "=A1*2"

    grid = _parse_workbook_grid(_workbook_bytes(populate))
    assert grid["sheets"][0]["rows"] == [[5, "=A1*2"]]


def test_trailing_empty_rows_and_columns_are_trimmed():
    def populate(wb):
        ws = wb.active
        ws["A1"] = "x"
        ws["E9"] = None  # touched but empty — must not widen the grid

    grid = _parse_workbook_grid(_workbook_bytes(populate))
    assert grid["sheets"][0]["rows"] == [["x"]]


def test_lying_dimension_record_does_not_truncate_the_grid():
    # Some writers emit a stale dimension ref (for example A1:A1) for a
    # populated sheet. Read-only openpyxl trusts it and iter_rows stops early,
    # which would silently drop cells from the preview; reset_dimensions()
    # forces a real scan.
    import io
    import re
    import zipfile

    def populate(wb):
        ws = wb.active
        ws["A1"] = "first"
        ws["C3"] = "later"

    data = _workbook_bytes(populate)
    src = zipfile.ZipFile(io.BytesIO(data))
    buf = io.BytesIO()
    replaced = False
    with zipfile.ZipFile(buf, "w") as out:
        for item in src.infolist():
            payload = src.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml" and b"<dimension" in payload:
                payload = re.sub(rb'<dimension ref="[^"]*"/>', b'<dimension ref="A1:A1"/>', payload)
                replaced = True
            out.writestr(item, payload)
    assert replaced, "fixture did not carry a dimension record to forge"

    grid = _parse_workbook_grid(buf.getvalue())
    rows = grid["sheets"][0]["rows"]
    assert len(rows) == 3
    assert rows[0][0] == "first"
    assert rows[2][2] == "later"


def test_row_cap_sets_truncated_flag_and_total():
    def populate(wb):
        ws = wb.active
        for i in range(_SHEET_MAX_ROWS + 40):
            ws.cell(row=i + 1, column=1, value=i)

    grid = _parse_workbook_grid(_workbook_bytes(populate))
    sheet = grid["sheets"][0]
    assert len(sheet["rows"]) == _SHEET_MAX_ROWS
    assert sheet["truncated_rows"] is True
    # No total is reported: any count derived from workbook geometry is
    # attacker-influenced (sparse row indices), so nothing iterates past the
    # cap to compute one.
    assert "total_rows" not in sheet


def test_column_cap_sets_truncated_flag():
    def populate(wb):
        ws = wb.active
        for c in range(_SHEET_MAX_COLS + 5):
            ws.cell(row=1, column=c + 1, value=c)

    grid = _parse_workbook_grid(_workbook_bytes(populate))
    sheet = grid["sheets"][0]
    assert len(sheet["rows"][0]) == _SHEET_MAX_COLS
    assert sheet["truncated_cols"] is True


def test_datetime_and_nonfinite_floats_serialize_json_safe():
    import datetime as dt
    import json

    def populate(wb):
        ws = wb.active
        ws["A1"] = dt.datetime(2026, 8, 16, 1, 30)
        ws["B1"] = dt.date(2026, 8, 16)

    grid = _parse_workbook_grid(_workbook_bytes(populate))
    row = grid["sheets"][0]["rows"][0]
    assert row[0] == "2026-08-16 01:30:00"
    # Excel stores dates as serial numbers; openpyxl rehydrates them as
    # midnight datetimes, so the round-tripped cell carries the time part.
    assert row[1] == "2026-08-16 00:00:00"
    # The serializer's date/time branches, hit directly.
    assert _sheet_cell_json(dt.date(2026, 8, 16)) == "2026-08-16"
    assert _sheet_cell_json(dt.time(9, 30)) == "09:30:00"
    # NaN/Infinity must not reach the JSON encoder as float — JSON.parse in the
    # browser rejects the JS-only tokens the stdlib encoder would emit.
    assert _sheet_cell_json(float("nan")) == "nan"
    assert _sheet_cell_json(float("inf")) == "inf"
    json.dumps(grid, allow_nan=False)  # whole payload must be strict-JSON safe


# --- Denials ----------------------------------------------------------------


def test_cell_text_is_credential_redacted():
    # Workbook text leaves the host through the dashboard -- the same egress
    # class as api_file_read, so the same redaction pass must apply.
    from kiro_crew.platform import redact_via_context

    sample = "aws key AKIA" + "IOSFODNN7EXAMPLE"
    assert redact_via_context(sample) != sample  # fixture must be redactable

    def populate(wb):
        ws = wb.active
        ws["A1"] = sample

    grid = _parse_workbook_grid(_workbook_bytes(populate))
    cell = grid["sheets"][0]["rows"][0][0]
    assert cell == redact_via_context(sample)
    assert "IOSFODNN7EXAMPLE" not in cell


def test_zip_expansion_cap_refuses_before_parse(tmp_path):
    import zipfile

    # Central directory declares a huge uncompressed member; the vet must
    # refuse before openpyxl inflates anything.
    f = tmp_path / "bomb.xlsx"
    with zipfile.ZipFile(f, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("big.bin", b"\x00" * (16 * 1024))
    with pytest.MonkeyPatch.context() as mp:
        mp.setattr("kiro_crew.dashboard.handlers.files._SHEET_MAX_EXPANDED_BYTES", 1024)
        with pytest.raises(_SheetRefusal) as exc:
            _load_sheet_payload(str(f))
    assert exc.value.status == 413
    assert "expands" in str(exc.value)


def test_zip_member_count_cap_refuses_before_parse(tmp_path):
    import zipfile

    f = tmp_path / "many.xlsx"
    with zipfile.ZipFile(f, "w") as z:
        for i in range(12):
            z.writestr(f"m{i}", b"x")
    with pytest.MonkeyPatch.context() as mp:
        mp.setattr("kiro_crew.dashboard.handlers.files._SHEET_MAX_MEMBERS", 10)
        with pytest.raises(_SheetRefusal) as exc:
            _load_sheet_payload(str(f))
    assert exc.value.status == 413


def test_eocd_declared_entry_count_refuses_before_zipfile_construction(tmp_path):
    import zipfile

    from kiro_crew.dashboard.handlers.files import _vet_zip_eocd

    # A crafted EOCD claiming hundreds of thousands of entries must be
    # refused from the raw bytes -- ZipFile construction itself would
    # materialize one ZipInfo per declared central-directory entry.
    f = tmp_path / "crafted.xlsx"
    with zipfile.ZipFile(f, "w") as z:
        z.writestr("m", b"x")
    data = bytearray(f.read_bytes())
    eocd = data.rfind(b"PK\x05\x06")
    data[eocd + 10 : eocd + 12] = (0xFFFE).to_bytes(2, "little")
    with pytest.raises(_SheetRefusal) as exc:
        _vet_zip_eocd(bytes(data))
    assert exc.value.status == 413
    assert exc.value.code == "workbook_expands_too_large"


def test_eocd_declared_directory_size_refuses_before_zipfile_construction(tmp_path):
    import zipfile

    from kiro_crew.dashboard.handlers.files import _vet_zip_eocd

    # A lying EOCD can under-declare the entry count while inflating the
    # directory byte size zipfile actually iterates by; both fields are capped.
    f = tmp_path / "crafted.xlsx"
    with zipfile.ZipFile(f, "w") as z:
        z.writestr("m", b"x")
    data = bytearray(f.read_bytes())
    eocd = data.rfind(b"PK\x05\x06")
    data[eocd + 12 : eocd + 16] = (0xFFFFFFF0).to_bytes(4, "little")
    with pytest.raises(_SheetRefusal) as exc:
        _vet_zip_eocd(bytes(data))
    assert exc.value.status == 413


def test_missing_eocd_refuses_as_not_a_spreadsheet():
    from kiro_crew.dashboard.handlers.files import _vet_zip_eocd

    with pytest.raises(_SheetRefusal) as exc:
        _vet_zip_eocd(b"PK\x03\x04" + b"\x00" * 64)
    assert exc.value.status == 415


def test_zip64_saturated_fields_refuse_outright(tmp_path):
    import zipfile

    from kiro_crew.dashboard.handlers.files import _vet_zip_eocd

    # Saturated classic EOCD fields declare an inventory orders of magnitude
    # past this endpoint's caps, so no ZIP64 record parse can change the
    # verdict -- the vet refuses without locating any further record, which
    # also removes the record-shadowing surface entirely.
    f = tmp_path / "crafted.xlsx"
    with zipfile.ZipFile(f, "w") as z:
        z.writestr("m", b"x")
    body = bytearray(f.read_bytes())
    eocd = body.rfind(b"PK\x05\x06")
    body[eocd + 10 : eocd + 12] = (0xFFFF).to_bytes(2, "little")
    data = b"PK\x06\x06" + b"\x00" * 52 + bytes(body)
    with pytest.raises(_SheetRefusal) as exc:
        _vet_zip_eocd(data)
    assert exc.value.status == 413
    assert exc.value.code == "workbook_expands_too_large"


def test_oversized_cell_text_is_truncated():
    from kiro_crew.dashboard.handlers.files import _SHEET_MAX_CELL_CHARS

    def populate(wb):
        wb.active["A1"] = "x" * (_SHEET_MAX_CELL_CHARS + 500)

    grid = _parse_workbook_grid(_workbook_bytes(populate))
    cell = grid["sheets"][0]["rows"][0][0]
    assert len(cell) == _SHEET_MAX_CELL_CHARS + 1  # truncated + ellipsis
    assert cell.endswith("…")


def test_workbook_text_budget_refuses_amplified_shared_strings():
    # A single shared string referenced by many cells amplifies far past the
    # archive-size caps; the cumulative text budget must refuse the preview.
    def populate(wb):
        ws = wb.active
        for r in range(1, 40):
            for c in range(1, 6):
                ws.cell(row=r, column=c, value="shared payload " * 10)

    with pytest.MonkeyPatch.context() as mp:
        mp.setattr("kiro_crew.dashboard.handlers.files._SHEET_MAX_TEXT_CHARS", 4000)
        with pytest.raises(_SheetRefusal) as exc:
            _parse_workbook_grid(_workbook_bytes(populate))
    assert exc.value.status == 413
    assert exc.value.code == "workbook_text_too_large"


@pytest.mark.asyncio
async def test_rejects_non_zip_magic(tmp_path, mock_sel):
    f = tmp_path / "fake.xlsx"
    f.write_bytes(b"not a zip at all")
    with patch("kiro_crew.dashboard.handlers._validate_dashboard_path", return_value=str(f)):
        async with TestClient(TestServer(_make_app())) as client:
            resp = await client.get(f"/api/file-sheet?path={f}")
            assert resp.status == 415


@pytest.mark.asyncio
async def test_zip_that_is_not_a_workbook_answers_422(tmp_path, mock_sel):
    import zipfile

    f = tmp_path / "junk.xlsx"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("hello.txt", "not a workbook")
    f.write_bytes(buf.getvalue())
    with patch("kiro_crew.dashboard.handlers._validate_dashboard_path", return_value=str(f)):
        async with TestClient(TestServer(_make_app())) as client:
            resp = await client.get(f"/api/file-sheet?path={f}")
            assert resp.status == 422
            assert "cannot parse" in (await resp.json())["error"]


@pytest.mark.asyncio
async def test_rejects_invalid_path(mock_sel):
    with patch("kiro_crew.dashboard.handlers._validate_dashboard_path", return_value=None):
        async with TestClient(TestServer(_make_app())) as client:
            resp = await client.get("/api/file-sheet?path=/etc/passwd")
            assert resp.status == 400


@pytest.mark.asyncio
async def test_rejects_sensitive_path(tmp_path):
    f = tmp_path / "creds.xlsx"
    f.write_bytes(b"PK\x03\x04junk")
    with patch("kiro_crew.sel.sel", return_value=MagicMock()), \
         patch("kiro_crew.security.is_sensitive_path", return_value=True), \
         patch("kiro_crew.dashboard.handlers._validate_dashboard_path", return_value=str(f)):
        async with TestClient(TestServer(_make_app())) as client:
            resp = await client.get(f"/api/file-sheet?path={f}")
            assert resp.status == 403


@pytest.mark.asyncio
async def test_missing_file_answers_404(tmp_path, mock_sel):
    ghost = tmp_path / "ghost.xlsx"
    with patch("kiro_crew.dashboard.handlers._validate_dashboard_path", return_value=str(ghost)):
        async with TestClient(TestServer(_make_app())) as client:
            resp = await client.get(f"/api/file-sheet?path={ghost}")
            assert resp.status == 404


@pytest.mark.asyncio
async def test_oversized_file_answers_413(tmp_path, mock_sel):
    f = tmp_path / "big.xlsx"
    f.write_bytes(b"PK\x03\x04" + b"\x00" * 16)
    with patch("kiro_crew.dashboard.handlers._validate_dashboard_path", return_value=str(f)), \
         patch("kiro_crew.dashboard.handlers.files._MAX_UPLOAD_BYTES", 8):
        async with TestClient(TestServer(_make_app())) as client:
            resp = await client.get(f"/api/file-sheet?path={f}")
            assert resp.status == 413


# --- Cancellation audit ------------------------------------------------------


@pytest.mark.asyncio
async def test_cancellation_during_parse_still_audits(tmp_path, mock_sel):
    """A CancelledError out of the worker parse must leave a SEL event.

    Gateway shutdown mid-parse cancels the awaiting handler; the file was
    already opened, so the access may not vanish from the audit trail. The
    handler logs a 'cancelled' outcome and re-raises.
    """
    f = tmp_path / "model.xlsx"
    f.write_bytes(_workbook_bytes(lambda wb: None))
    with patch("kiro_crew.dashboard.handlers._validate_dashboard_path", return_value=str(f)), \
         patch(
             "kiro_crew.dashboard.handlers.files.asyncio.to_thread",
             side_effect=asyncio.CancelledError,
         ):
        request = MagicMock()
        request.query = {"path": str(f)}
        with pytest.raises(asyncio.CancelledError):
            await api_file_sheet(request)
    outcomes = [
        call.kwargs.get("outcome")
        for call in mock_sel.log_tool_invocation.call_args_list
    ]
    assert "cancelled" in outcomes
